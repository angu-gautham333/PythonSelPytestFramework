import openpyxl

class HomePageData:

    test_HomePage_data = [{"firstname":"Anguvel","email":"Gautham@test.com","password":"password2","gender":"Male"},
                            {"firstname":"Giii","email":"Gautham@test.com","password":"password3","gender":"Female"},]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("D:\\python testing\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
